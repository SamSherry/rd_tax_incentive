
# coding: utf-8

# In[52]:

# Combine metals and mining properties data into a single file
import os
import sys
import pandas as pd
import numpy as np
import openpyxl
import xlrd
from xlrd import open_workbook, cellname
MY_DIR = '/media/sf_Shared/resources/'
sys.path.append(MY_DIR)
frames = []
for f in os.listdir(MY_DIR):
        if f.endswith(".xls") and f != "combined.xls":
            print(f)
            wb = xlrd.open_workbook(MY_DIR + f)
            for sheet in wb.sheets():
                if sheet.nrows > 0:
                    if sheet.cell(1,0).value == 'Metals & Mining Properties':
                        print sheet.cell(0,0).value
                        print sheet.name
                        frame = pd.read_excel(MY_DIR + f, sheet.name, skiprows=int(12), skip_footer=int(6))
                        company = sheet.cell(0,0).value
                        frame['filename'] = f
                        frame['company'] = company
                        frames.append(frame)
                else:
                    print sheet.name
                
properties = pd.concat(frames)


# In[53]:

properties.info()


# In[54]:

properties['company'].value_counts()


# In[55]:

properties['company'].nunique()


# In[56]:

properties['Property'].nunique()


# In[57]:

properties['Property ID'].nunique()


# In[58]:

properties['Primary Commodity'].value_counts()


# In[59]:

# Commodity names
commodities = pd.Series(properties['Commodity(s)'])
commodities = commodities.str.split(',', expand=True)


# In[60]:

commodities[0].value_counts()


# In[61]:

commodities.info()


# In[62]:

# Rename columns
commodities.rename(columns={0:'commodity1',1:'commodity2',2:'commodity3',3:'commodity4',4:'commodity5',5:'commodity6',
                           6:'commodity7',7:'commodity8',8:'commodity9',9:'commodity10',10:'commodity11',11:'commodity12',
                           12:'commodity13',13:'commodity14',14:'commodity15',15:'commodity16',16:'commodity17'},inplace=True)


# In[63]:

commodities.head()


# In[64]:

commodities = commodities.reset_index()


# In[65]:

commodities.index


# In[66]:

commodities.drop('index',axis=1,inplace=True)


# In[67]:

commodities.head()


# In[ ]:

commodities.to_csv('commodities.csv')


# In[68]:

properties = properties.reset_index()


# In[69]:

properties.index


# In[70]:

properties.head()


# In[71]:

properties_v2 = pd.concat([properties,commodities],axis=1)
properties_v2.info()


# In[72]:

properties_v2.head()


# In[74]:

commod_mismatch = properties_v2[properties_v2['Primary Commodity']!=properties_v2['commodity1']]
commod_mismatch


# In[77]:

properties_v2.drop('commodity1',axis=1,inplace=True)


# In[ ]:

properties_v2.drop("Commodity(s)",axis=1,inplace=True)


# In[ ]:

# Remove surplus characters from column names and check results
properties_v2.rename(columns = lambda x: x.replace("\n"," ").replace(u"\xb1","").replace(u"\xb2",""),inplace=True)


# In[ ]:

commod_mismatch.to_excel('mismatch.xls')


# In[86]:

# Primary commodity list
primarycommlist = pd.Series(properties_v2['Primary Commodity'].unique())
primarycommlist
primarycommlist.to_csv('/media/sf_Shared/commlist.csv')


# In[87]:

# Commodity IDs
comm_index = pd.read_csv('/media/sf_Shared/commlist.csv', header=None)
comm_index.rename(columns={0:'Commodity ID',1:'Primary Commodity'},inplace=True)
comm_index


# In[88]:

# Merge commodity IDs with property table
properties_merged = pd.merge(properties_v2,comm_index,how='inner',on='Primary Commodity')


# In[102]:

properties_merged.columns


# In[103]:

# Filenames
ticker = pd.Series(properties_merged['filename'])
ticker = ticker.str.replace('.xls',' ').str.strip()
ticker.nunique()


# In[106]:

ticker = pd.DataFrame(ticker)


# In[108]:

ticker.rename(columns = {'filename':'Ticker'}, inplace=True)


# In[109]:

ticker


# In[111]:

properties_ticker = pd.concat([properties_merged, ticker],axis=1)
properties_ticker.columns


# In[112]:

properties_ticker.rename(columns = {'company':'Company', 'Unnamed: 16':'Units',
                                    u'Total In-situ Value¹ ($M)':'Value'}, inplace=True)


# In[114]:

properties_ticker.head()


# In[ ]:

# Eliminate duplicate Property IDs


# In[121]:

properties_ticker.rename(columns = {'index':'Property Index'}, inplace=True)


# In[122]:

properties_ticker.index


# In[123]:

properties_ticker.info()


# In[129]:

properties_ticker['Property'].value_counts() # legitimate duplicates - keep


# In[138]:

# Duplicates - these mines appear more than once in the database because they have more than one owner.
properties_ticker['Property ID'].nunique()


# In[139]:

properties_ticker.index


# In[135]:

properties_unique = properties_ticker.drop_duplicates('Property ID', keep='first')


# In[140]:

properties_unique.info()


# In[141]:

properties_unique.to_excel('/media/sf_Shared/properties_unique.xlsx')

