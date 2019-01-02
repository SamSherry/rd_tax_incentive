
# coding: utf-8

# In[1]:

# Concatenate resources and reserves data - need to install xlwt
import os
import sys
import pandas as pd
import numpy as np
import openpyxl
import xlrd
from xlrd import open_workbook, cellname
MY_DIR = '/media/sf_Shared/resources/'
sys.path.append(MY_DIR)
frames = []
for f in os.listdir(MY_DIR):
        if f.endswith(".xls") and f != "combined.xls":
            print(f)
            wb = xlrd.open_workbook(MY_DIR + f)
            for sheet in wb.sheets():
                if sheet.nrows > 0:
                    if sheet.cell(1,0).value == 'Ownership':
                        if sheet.name == 'Current Ownership':
                            print sheet.cell(0,0).value
                            print sheet.name
                            frame = pd.read_excel(MY_DIR + f, sheet.name, skiprows=int(4))
                            project = sheet.cell(0,0).value
                            print sheet.cell(1,0).value
                            frame['filename'] = f
                            frame['project'] = project
                            frames.append(frame)
                else:
                    print sheet.name
                    
current_ownership = pd.concat(frames)


# In[2]:

current_ownership['project'].nunique()


# In[3]:

current_ownership['filename'].nunique()


# In[4]:

current_ownership.info()


# In[5]:

current_ownership.head()


# In[6]:

current_ownership.tail()


# In[7]:

current_ownership.columns


# In[8]:

current_ownership.to_excel('/media/sf_Shared/curr_ownership.xlsx')


# In[ ]:



