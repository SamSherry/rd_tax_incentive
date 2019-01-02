
# coding: utf-8

# In[1]:

# Concatenate resources and reserves data - need to install xlwt
import os
import sys
import pandas as pd
import numpy as np
import openpyxl
import xlrd
from xlrd import open_workbook, cellname
MY_DIR = '/media/sf_Shared/SNL_Resource_and_Reserve_Information'
sys.path.append(MY_DIR)
frames = []
for f in os.listdir(MY_DIR):
        if f.endswith(".xls") and f != "combined.xls":
            print(f)
            wb = xlrd.open_workbook(MY_DIR + f)
            for sheet in wb.sheets():
                if sheet.nrows > 0:
                    if sheet.cell(1,0).value == 'Reserves & Resources':
                        print sheet.cell(0,0).value
                        print sheet.name
                        path = MY_DIR + f
                        print path
                        print sheet.cell(1,0).value
                        frame['filename'] = f
                        frame['path'] = path
                        frames.append(frame)
                else:
                    print sheet.name
                    
filenames = pd.concat(frames)


# In[ ]:



