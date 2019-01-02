
# coding: utf-8

# In[1]:

# Concatenate resources and reserves data - need to install xlwt
import os
import sys
import pandas as pd
import numpy as np
import openpyxl
import xlrd
from xlrd import open_workbook, cellname
MY_DIR = '/media/sf_Shared/resources/'
sys.path.append(MY_DIR)
frames = []
for f in os.listdir(MY_DIR):
        if f.endswith(".xls") and f != "combined.xls":
            print(f)
            wb = xlrd.open_workbook(MY_DIR + f)
            for sheet in wb.sheets():
                if sheet.nrows > 0:
                    if sheet.cell(1,0).value == 'Ownership':
                        if sheet.name == 'Historical Ownership':
                            print sheet.cell(0,0).value
                            print sheet.name
                            frame = pd.read_excel(MY_DIR + f, sheet.name, skiprows=int(3))
                            project = sheet.cell(0,0).value
                            print sheet.cell(1,0).value
                            frame['filename'] = f
                            frame['project'] = project
                            frames.append(frame)
                else:
                    print sheet.name
                    
historic_ownership = pd.concat(frames)


# In[2]:

historic_ownership['project'].nunique()


# In[3]:

historic_ownership['filename'].nunique()


# In[4]:

historic_ownership.info()


# In[5]:

historic_ownership.head()


# In[6]:

historic_ownership.tail()


# In[7]:

historic_ownership.columns


# In[8]:

historic_ownership.to_excel('/media/sf_Shared/historic_ownership.xlsx')

