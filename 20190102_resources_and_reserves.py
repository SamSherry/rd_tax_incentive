# Concatenate resources and reserves data - need to install xlwt
import os
import sys
import pandas as pd
import numpy as np
import openpyxl
import xlrd
from xlrd import open_workbook, cellname
MY_DIR = '/media/sf_Shared/resources/'
sys.path.append(MY_DIR)
frames = []
for f in os.listdir(MY_DIR):
        if f.endswith(".xls") and f != "combined.xls":
            print(f)
            wb = xlrd.open_workbook(MY_DIR + f)
            for sheet in wb.sheets():
                if sheet.nrows > 0:
                    if sheet.cell(1,0).value == 'Reserves & Resources':
                        print sheet.cell(0,0).value
                        print sheet.name
                        path = MY_DIR + f
                        print MY_DIR + f
                        frame = pd.read_excel(MY_DIR + f, sheet.name, skiprows=int(4))
                        project = sheet.cell(0,0).value
                        print sheet.cell(1,0).value
                        frame['filename'] = f
                        frame['project'] = project
                        frames.append(frame)
                else:
                    print sheet.name
                    
combined = pd.concat(frames)
combined['project'].nunique()
combined['filename'].nunique()

# In[4]:

combined.info()

# In[5]:

combined.head()


# In[6]:

combined.tail()


# In[7]:

combined.columns


# In[8]:

# Drop columns that aren't needed
combined.drop(['Unnamed: 6', 'Unnamed: 7', 'Unnamed: 8', 'Unnamed: 9', 'Unnamed: 10',
               'Unnamed: 11', 'Unnamed: 12', 'Unnamed: 13', 'Unnamed: 14', 'Unnamed: 15'], axis=1, inplace=True)


# In[9]:

# Rename columns
combined.rename(columns = {'As of Date':'Date','Unnamed: 1':'Disclosure','Unnamed: 2':'Certainty','Unnamed: 3':'Classification',
                           'Unnamed: 4':'CutoffGrade'}, inplace=True)


# In[10]:

# Remove surplus characters from column names and check results
combined.rename(columns = lambda x: x.replace("\n","").replace(u"\xb2","2").replace(u"\xb3","3"), inplace=True)


# In[11]:

# Save new column names to a dataframe
colnames = pd.DataFrame(combined.columns)
colnames
# Export column names to CSV
colnames.to_csv('/media/sf_Shared/colnames_new.csv')
# Export column names to Excel
colnames.to_excel('/media/sf_Shared/columns.xlsx')


# In[12]:

combined['Date'].value_counts()


# In[13]:

type(combined['Date'])


# In[14]:

combined = combined.reset_index()


# In[15]:

combined.index


# In[16]:

# Deleting rows that contain no data
bad_rows = combined['Date'].str.startswith('A', na=False)
combined.loc[bad_rows, 'Date'] = np.nan


# In[17]:

bad_rows = combined['Date'].str.startswith('I', na=False)
combined.loc[bad_rows, 'Date'] = np.nan


# In[18]:

bad_rows = combined['Date'].str.startswith('O', na=False)
combined.loc[bad_rows, 'Date'] = np.nan


# In[19]:

bad_rows = combined['Date'].str.startswith('P', na=False)
combined.loc[bad_rows, 'Date'] = np.nan


# In[20]:

bad_rows = combined['Date'].str.startswith('R', na=False)
combined.loc[bad_rows, 'Date'] = np.nan


# In[21]:

bad_rows = combined['Date'].str.startswith('S', na=False)
combined.loc[bad_rows, 'Date'] = np.nan


# In[22]:

bad_rows = combined['Date'].str.startswith('T', na=False)
combined.loc[bad_rows, 'Date'] = np.nan


# In[23]:

bad_rows = combined['Date'].str.contains('Inclusion', na=False)
combined.loc[bad_rows, 'Date'] = np.nan


# In[24]:

bad_rows = combined['Date'].str.contains('reserves', na=False)
combined.loc[bad_rows, 'Date'] = np.nan


# In[25]:

combined2 = combined[combined['Date'].notnull()]


# In[26]:

combined2 = combined2.reset_index()
combined2 = combined2.drop('level_0', axis=1)


# In[103]:

combined2.index


# In[27]:

combined2.info()


# In[28]:

# Datetime index
from datetime import datetime
datetime_index = pd.DatetimeIndex(combined2['Date'])
type(datetime_index)


# In[29]:

datetime_index.min()


# In[30]:

unique_dates = pd.DataFrame(datetime_index.value_counts())
unique_dates


# In[31]:

unique_dates.to_excel('/media/sf_Shared/dates.xlsx')


# In[32]:

# Extract primary commodities
primary_comm = pd.Series(combined2['project'], name='Primary Commodity').str.split('(').str[-1]
primary_comm = primary_comm.str.split(':').str[-1].str.replace(')',' ').str.strip()
primary_comm.unique()


# In[33]:

# Export primary commodities list to Excel
primary_comm.value_counts()
commod_counts = pd.DataFrame(primary_comm.value_counts())
commod_counts.to_excel('/media/sf_Shared/commod_counts.xlsx')


# In[35]:

# Extract project names
project_name = pd.Series(combined2['project'], name='Property')
project_name = project_name.str.split('(').str[0].str.strip()
project_name.nunique()


# In[36]:

project_name.value_counts()


# In[37]:

# Export project names to Excel
project_names = pd.DataFrame(project_name.value_counts())
project_names.to_excel('/media/sf_Shared/project_list.xlsx')


# In[38]:

# Concatenate into new dataframe
combined3 = pd.concat([combined2,project_name,primary_comm],axis=1)


# In[39]:

combined3['Disclosure'].value_counts()


# In[40]:

disclosures = pd.DataFrame(combined3['Disclosure'].value_counts())
disclosures.to_excel('/media/sf_Shared/disclosures.xlsx')


# In[41]:

filenames = pd.DataFrame(combined3['filename'].value_counts())
filenames.to_excel('/media/sf_Shared/filenames.xlsx')


# In[42]:

pd.Series(combined3.columns)


# In[43]:

# Save new column names to a dataframe
colnames = pd.DataFrame(combined3.columns)
# Export column names to Excel
colnames.to_excel('/media/sf_Shared/columns.xlsx')


# In[44]:

# Set datetime index
combined3['Date'] = datetime_index
combined3.set_index('Date',inplace=True)


# In[45]:

# Extract data you need
resources_reserves = combined3[(combined3['Disclosure'].str.contains('Total Reserves & Resources', na=False))]
resources = combined3[(combined3['Disclosure'].str.contains('Total Resources Exclusive of Reserves', na=False))]
reserves = combined3[(combined3['Disclosure'].str.contains('Total Reserves', na=False))]


# In[46]:

resources_reserves.info()


# In[47]:

resources.info()


# In[48]:

reserves.info()


# In[49]:

resources_reserves.to_csv('/media/sf_Shared/resources_reserves.csv')


# In[50]:

resources.to_csv('/media/sf_Shared/resources.csv')


# In[51]:
reserves.to_csv('/media/sf_Shared/reserves.csv')
